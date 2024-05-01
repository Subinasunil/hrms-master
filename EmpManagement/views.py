from django.conf import settings
from django.shortcuts import render
import datetime
from Core.models import Document_type
from .models import (emp_family,Emp_Documents,EmpJobHistory,EmpLeaveRequest,EmpQualification,
                     emp_master,Emp_CustomField,EmpFamily_CustomField,EmpJobHistory_CustomField,
                     EmpQualification_CustomField,EmpDocuments_CustomField,LanguageSkill,MarketingSkill,ProgrammingLanguageSkill,EmployeeSkill)
from .serializer import (Emp_qf_Serializer,EmpFamSerializer,EmpSerializer,
                         EmpJobHistorySerializer,EmpLeaveRequestSerializer,DocumentSerializer,EmpBulkUploadSerializer,CustomFieldSerializer,
                         EmpFam_CustomFieldSerializer,EmpJobHistory_Udf_Serializer,Emp_qf_udf_Serializer,EmpDocuments_Udf_Serializer,
                         DocBulkuploadSerializer,LanguageSkillSerializer,MarketingSkillSerializer,ProgrammingLanguageSkillSerializer,EmployeeSkillSerializer,
                         LanguageBlkupldSerializer,MarketingBlkupldSerializer,ProLangBlkupldSerializer)
from rest_framework.decorators import action,api_view
from rest_framework import viewsets,filters,parsers
from rest_framework.response import Response
from rest_framework import status,generics,viewsets,permissions
# from rest_framework.authentication import SessionAuthentication,TokenAuthentication
from rest_framework.permissions import IsAuthenticated,AllowAny,IsAuthenticatedOrReadOnly,IsAdminUser
# from django.views.decorators.csrf import csrf_exempt
from rest_framework.parsers import JSONParser
# from rest_framework.decorators import api_view
# from rest_framework.response import Response
from django.http import FileResponse,HttpResponse,JsonResponse

from openpyxl import Workbook
from rest_framework.parsers import MultiPartParser, FormParser
from .resource import EmployeeResource,EmpResource_Export,EmpCustomFieldResource,DocumentResource,LanguageSkillResource,MarketingSkillResource,ProLangSkillResource
import tablib
import pandas as pd
import openpyxl
import xlsxwriter
import io
import os
from django.db import transaction
from django.core.exceptions import ValidationError
from import_export.results import RowResult
from tablib import Dataset
from io import BytesIO
from rest_framework.permissions import IsAuthenticated
# from .permissions import CanExportDataPermission
from rest_framework.permissions import IsAdminUser
from django.contrib.auth.decorators import permission_required
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from openpyxl.styles import PatternFill,Alignment
from django.core.files import File 

from django.http import HttpResponse
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter,landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib.units import inch
import requests
import json
from collections import defaultdict

# from django. utils. timezone import timedelta
# from django.utils import timezone
# from UserManagement.models import CustomUser
# Create your views here.
#EMPLOYEE CRUD
class EmpViewSet(viewsets.ModelViewSet):
    queryset = emp_master.objects.all()
    serializer_class = EmpSerializer
    permission_classes = [IsAuthenticated]
    parser_classes = (MultiPartParser, FormParser)

   


    @action(detail=True, methods=['get'])
    def emp_family(self, request, pk=None):
        employee = self.get_object()
        families = employee.emp_family.all()  
        serializer = EmpFamSerializer(families, many=True)
        return Response(serializer.data)
    @action(detail=True, methods=['get'])
    def emp_qualification(self, request, pk=None):
        employee = self.get_object()
        qualification = employee.emp_qualification.all()  
        serializer = Emp_qf_Serializer(qualification, many=True)
        return Response(serializer.data)
    @action(detail=True, methods=['get'])
    def emp_job_history(self, request, pk=None):
        employee = self.get_object()
        job_history = employee.emp_job_history.all()  
        serializer = EmpJobHistorySerializer(job_history, many=True)
        return Response(serializer.data)
    @action(detail=True, methods=['get'])
    def emp_documents(self, request, pk=None):
        employee = self.get_object()
        Emp_Documents = employee.emp_documents.all()  
        serializer = DocumentSerializer(Emp_Documents, many=True)
        return Response(serializer.data)
    
    @action(detail=True, methods=['get'])
    def emp_udf(self, request, pk=None):
        udf = self.get_object()
        userdefinedfields = udf.custom_fields.all()  
        serializer = CustomFieldSerializer(userdefinedfields, many=True)
        return Response(serializer.data)


    def get_queryset(self):
        user = self.request.user
        if user.is_authenticated:
            if user.is_ess:  # If user is an ESS, they can only access their own employee information
                return emp_master.objects.filter(created_by=user)
            else:
                return emp_master.objects.all()  # Other users can access all employee information
    
        return emp_master.objects.none()  # Return an empty queryset for unauthenticated users

    


    @action(detail=False, methods=['get'],permission_classes=[IsAdminUser])
    def export_employee_data(self, request):
        if not request.user.is_superuser:
            return Response({"error": "You do not have permission to access this resource."}, status=status.HTTP_403_FORBIDDEN)
        queryset = emp_master.objects.all()
        resource = EmpResource_Export()
        dataset = resource.export(queryset)

        # Create an Excel workbook
        wb = Workbook()
        ws = wb.active

        # Define default cell formats
        default_fill = PatternFill(start_color='FFC0CB', end_color='FFC0CB', fill_type='solid')  # pink background color
        default_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Apply default styles to headings
        for col_num, field_name in enumerate(dataset.headers):
            ws.cell(row=1, column=col_num + 1, value=field_name)
            ws.cell(row=1, column=col_num + 1).fill = default_fill
            ws.cell(row=1, column=col_num + 1).alignment = default_alignment

        # Write data to the Excel file
        for row_num, row in enumerate(dataset, start=2):
            for col_num, value in enumerate(row, start=1):
                ws.cell(row=row_num, column=col_num, value=value)
                ws.cell(row=row_num, column=col_num).alignment = default_alignment

        # Auto-size columns
        for col in ws.columns:
            max_length = 0
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[col[0].column_letter].width = adjusted_width

        # Save the workbook to an in-memory buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        # Prepare response
        response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="exported_data.xlsx"'
        return response

    @action(detail=False, methods=['get'])
    def emp_pdf_report(self, request, *args, **kwargs):
        # Query employee details from the database
        employees = emp_master.objects.all()

        # Create a response object
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="employee_report.pdf"'

        # Create a PDF document
        pdf = SimpleDocTemplate(response, pagesize=letter)
        
        # Create a table to hold employee details
        data = [["ID", "FirstName", "LastName", "Gender", "DOB", "Email", "Mobile Number", "Country", "State", "City", "Permanent Address", "Present Address", "Status", "Hired Date", "Religion", "Blood Group", "Nationality", "Marital Status", "Father Name", "Mother Name", "Posting Location","Created At", "Created By", "Updated At", "Updated By","Active","OT Applicable", "Company", "Branch", "Department", "Designation", "Category"]]
        
        for employee in employees:
            data.append([
                employee.emp_code,
                employee.emp_first_name,
                employee.emp_last_name,
                employee.get_emp_gender_display() if employee.emp_gender else "",  # Get display value for choices field
                employee.emp_date_of_birth,
                employee.emp_personal_email,
                employee.emp_mobile_number_1,
                employee.emp_country_id.country_name if employee.emp_country_id else "",  # Assuming 'name' is the field for country name
                employee.emp_state_id.state_name if employee.emp_state_id else "",  # Assuming 'name' is the field for state name
                employee.emp_city,
                employee.emp_permenent_address,
                employee.emp_present_address,
                "Active" if employee.emp_status else "Inactive",
                employee.emp_hired_date,
                employee.emp_relegion,
                employee.emp_blood_group,
                employee.emp_nationality_id,
                employee.get_emp_marital_status_display() if employee.emp_marital_status else "",  # Get display value for choices field
                employee.emp_father_name,
                employee.emp_mother_name,
                employee.emp_posting_location,
                employee.created_at,
                employee.created_by.username if employee.created_by else "",  # Assuming 'username' is the field for username
                employee.updated_at,
                employee.updated_by.username if employee.updated_by else "",  # Assuming 'username' is the field for username
                "Active" if employee.is_active else "Inactive",
                "Yes" if employee.epm_ot_applicable else "No",
                employee.emp_company_id.cmpny_name if employee.emp_company_id else "",  # Assuming 'name' is the field for company name
                employee.emp_branch_id.branch_name if employee.emp_branch_id else "",  # Assuming 'name' is the field for branch name
                employee.emp_dept_id.dept_name if employee.emp_dept_id else "",  # Assuming 'name' is the field for department name
                employee.emp_desgntn_id.job_title if employee.emp_desgntn_id else "",  # Assuming 'name' is the field for designation name
                employee.emp_ctgry_id.catogary_title if employee.emp_ctgry_id else "",  # Assuming 'name' is the field for category name
            ])

        
        # Calculate column widths based on content
        col_widths = [1 * inch] * len(data[0])  # Initialize with default width (1.2 inch per column)
        for row in data:
            for i, cell in enumerate(row):
                col_widths[i] = max(col_widths[i], len(str(cell)) * 0.10 * inch)  # Adjust width based on content length

        # Calculate total table width
        total_width = sum(col_widths)

        # Set page size to accommodate the table width
        pdf.pagesize = (total_width, letter[1])  # Use the total width calculated and keep the height same as letter page height



       # Create the table with adjusted column widths and enable text wrapping
        table = Table(data, colWidths=col_widths, repeatRows=1)
        table.setStyle(TableStyle([('WORDWRAP', (0, 0), (-1, -1), True)]))

        # Add style to the table
        style = TableStyle([('BACKGROUND', (0, 0), (-1, 0), colors.gray),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),  # Vertical alignment
                    ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),  # Text color
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),  # Header text color
                    ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),  # Font
                    ('FONT_SIZE', (0, 0), (-1, -1), 10),  # Font size
                    ('WORD_WRAP', (0, 0), (-1, -1), True)])  # Enable word wrap

        table.setStyle(style)

        # Add the table to the PDF document
        elements = [table]
        pdf.build(elements)

        return response

    @action(detail=False, methods=['get'])
    def emp_excel_report(self, request, *args, **kwargs):
    # Query employee details from the database
        employees = emp_master.objects.all()

        # Create a response object
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="employee_report.xlsx"'

        # Create an Excel workbook and add a worksheet
        workbook = xlsxwriter.Workbook(response)
        worksheet = workbook.add_worksheet()

        # Add column headers
        headers = ["Employee Code", "FirstName", "LastName", "Gender", "DOB", "Email", "Mobile Number","Mobile Number2", "Country", "State", "City", "Permanent Address", "Present Address", "Status", "Hired Date","Active Date", "Religion", "Blood Group", "Nationality", "Marital Status", "Father Name", "Mother Name", "Posting Location", "Active", "OT Applicable", "Company", "Branch", "Department", "Designation", "Category"]
        for col, header in enumerate(headers):
            worksheet.write(0, col, header)
        worksheet.set_column('A:AD', 15)
        # Write data rows
        date_format = workbook.add_format({'num_format': 'dd-mm-yyyy'})
        for row, employee in enumerate(employees, start=1):
            worksheet.write(row, 0, employee.emp_code)
            worksheet.write(row, 1, employee.emp_first_name)
            worksheet.write(row, 2, employee.emp_last_name)
            worksheet.write(row, 3, employee.get_emp_gender_display() if employee.emp_gender else "")
            worksheet.write(row, 4, employee.emp_date_of_birth,date_format)
            worksheet.write(row, 5, employee.emp_personal_email)
            worksheet.write(row, 6, employee.emp_mobile_number_1)
            worksheet.write(row, 7, employee.emp_mobile_number_2)
            worksheet.write(row, 8, employee.emp_country_id.country_name if employee.emp_country_id else "",)
            worksheet.write(row, 9, employee.emp_state_id.state_name if employee.emp_state_id else "",)
            worksheet.write(row, 10, employee.emp_city)
            worksheet.write(row, 11, employee.emp_permenent_address)
            worksheet.write(row, 12, employee.emp_present_address)
            worksheet.write(row, 13, employee.emp_status)
            worksheet.write(row, 14, employee.emp_hired_date,date_format)
            worksheet.write(row, 15, employee.emp_active_date,date_format)
            worksheet.write(row, 16, employee.emp_relegion)
            # worksheet.write(row, 17, employee.emp_profile_pic)
            worksheet.write(row, 17, employee.emp_blood_group)
            worksheet.write(row, 18, employee.emp_nationality_id.N_name if employee.emp_nationality_id else "",)
            worksheet.write(row, 19, employee.emp_marital_status)
            worksheet.write(row, 20, employee.emp_father_name)
            worksheet.write(row, 21, employee.emp_mother_name)
            worksheet.write(row, 22, employee.emp_posting_location)
            worksheet.write(row, 23, employee.is_active)
            worksheet.write(row, 24, employee.epm_ot_applicable)
            worksheet.write(row, 25, employee.emp_company_id.cmpny_name if employee.emp_company_id else "",)
            worksheet.write(row, 26, employee.emp_branch_id.branch_name if employee.emp_branch_id else "",)
            worksheet.write(row, 27, employee.emp_dept_id.dept_name if employee.emp_dept_id else "",)
            worksheet.write(row, 28, employee.emp_desgntn_id.job_title if employee.emp_desgntn_id else "",)
            worksheet.write(row, 29, employee.emp_ctgry_id.catogary_title if employee.emp_ctgry_id else "")
            # worksheet.write(row, 31, employee.emp_languages)
           
            
        # worksheet.protect()
        # Close the workbook
        workbook.close()
        return response
    @action(detail=False, methods=['get'])
    def select_fields(self, request, *args, **kwargs):
        # Get all available fields
        available_fields = ["emp_code", "emp_first_name", "emp_last_name", "emp_gender", "emp_date_of_birth",
                            "emp_personal_email", "emp_mobile_number_1", "emp_mobile_number_2", "emp_country_id",
                            "emp_state_id", "emp_city", "emp_permenent_address", "emp_present_address",
                            "emp_status", "emp_hired_date", "emp_active_date", "emp_relegion",
                            "emp_blood_group", "emp_nationality_id", "emp_marital_status", "emp_father_name",
                            "emp_mother_name", "emp_posting_location", "is_active", "epm_ot_applicable",
                            "emp_company_id", "emp_branch_id", "emp_dept_id", "emp_desgntn_id", "emp_ctgry_id"]
        
        # Pass available fields to the template
        return render(request, 'select.html', {'available_fields': available_fields})


    @action(detail=False, methods=['post'])
    def emp_select_report(self, request, *args, **kwargs):
        # Get the list of fields selected by the user from the request data
        fields_to_include = request.POST.getlist('fields', [])

        # If no fields are selected, return a bad request response
        if not fields_to_include:
            return HttpResponse("No fields selected", status=400)

        # Mapping of database field names to custom column headings
        column_headings = {
            "emp_code": "Employee Code",
            "emp_first_name": "First Name",
            "emp_last_name": "Last Name",
            "emp_gender": "Gender",
            "emp_date_of_birth": "Date of Birth",
            "emp_personal_email": "Email",
            "emp_mobile_number_1": "Mobile Number",
            "emp_mobile_number_2": "Mobile Number2",
            "emp_country_id__country_name": "Country",
            "emp_state_id__state_name": "State",
            "emp_city": "City",
            "emp_permenent_address": "Permanent Address",
            "emp_present_address": "Present Address",
            "emp_status": "Status",
            "emp_hired_date": "Hired Date",
            "emp_active_date": "Active Date",
            "emp_relegion": "Religion",
            "emp_blood_group": "Blood Group",
            "emp_nationality_id": "Nationality",
            "emp_marital_status": "Marital Status",
            "emp_father_name": "Father Name",
            "emp_mother_name": "Mother Name",
            "emp_posting_location": "Posting Location",
            "is_active": "Active",
            "epm_ot_applicable": "OT Applicable",
            "emp_company_id": "Company",
            "emp_branch_id": "Branch",
            "emp_dept_id": "Department",
            "emp_desgntn_id": "Designation",
            "emp_ctgry_id": "Category",
            # Add more mappings as needed
        }

        # Query employee details from the database
        employees = emp_master.objects.all()
  
        # Create a response object
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="employee_report.xlsx"'

        # Create an Excel workbook and add a worksheet
        workbook = xlsxwriter.Workbook(response)
        worksheet = workbook.add_worksheet()

        # Write column headers for the selected fields
        headers = [column_headings.get(field, field) for field in fields_to_include]
        for col, header in enumerate(headers):
            worksheet.write(0, col, header)

        worksheet.set_column('A:AD', 15)

        # Write data rows for the selected fields
        date_format = workbook.add_format({'num_format': 'dd-mm-yyyy'})
        for row, employee in enumerate(employees, start=1):
            for col, field_name in enumerate(fields_to_include):
                field_value = getattr(employee, field_name)
                if field_name == 'emp_country_id':
                    field_value = employee.emp_country_id.country_name if employee.emp_country_id else ""
                elif field_name == 'emp_state_id':
                    field_value = employee.emp_state_id.state_name if employee.emp_state_id else ''
                elif field_name == 'emp_company_id':
                    field_value = employee.emp_company_id.cmpny_name if employee.emp_company_id else ''
                elif field_name == 'emp_branch_id':
                    field_value = employee.emp_branch_id.branch_name if employee.emp_branch_id else ''
                elif field_name == 'emp_dept_id':
                    field_value = employee.emp_dept_id.dept_name if employee.emp_dept_id else ''
                elif field_name == "emp_desgntn_id": 
                    field_value = employee.emp_desgntn_id.job_title if employee.emp_desgntn_id else ""
                elif field_name == "emp_ctgry_id": 
                    field_value = employee.emp_ctgry_id.catogary_title if employee.emp_ctgry_id else ""
                
                elif isinstance(field_value, datetime.date):
                    worksheet.write(row, col, field_value.strftime('%d-%m-%Y'), date_format)
                else:
                    worksheet.write(row, col, field_value)

        # Close the workbook
        workbook.close()
        return response
            
          
    # @action(detail=False, methods=['get'])
    # def emp_select_report(self, request, *args, **kwargs):
    #         # Query employee details from the database
    #         employees = emp_master.objects.all()

    #         # Fields to include in the report
    #         fields_to_include = request.GET.getlist('fields', [])  # Get list of fields from query parameters
            
    #         # If no fields are specified, include all fields
    #         if not fields_to_include:
    #             fields_to_include = ["emp_code", "emp_first_name", "emp_last_name", "emp_gender", "emp_date_of_birth",
    #                                 "emp_personal_email", "emp_mobile_number_1", "emp_mobile_number_2", "emp_country_id",
    #                                 "emp_state_id", "emp_city", "emp_permenent_address", "emp_present_address",
    #                                 "emp_status", "emp_hired_date", "emp_active_date", "emp_relegion",
    #                                 "emp_blood_group", "emp_nationality_id", "emp_marital_status", "emp_father_name",
    #                                 "emp_mother_name", "emp_posting_location", "is_active", "epm_ot_applicable",
    #                                 "emp_company_id", "emp_branch_id", "emp_dept_id", "emp_desgntn_id", "emp_ctgry_id"]
    #         else:
    #             # Filter out invalid fields
    #             valid_fields = [field for field in fields_to_include if field in ["emp_code", "emp_first_name", "emp_last_name", "emp_gender", "emp_date_of_birth",
    #                                                                             "emp_personal_email", "emp_mobile_number_1", "emp_mobile_number_2", "emp_country_id",
    #                                                                             "emp_state_id", "emp_city", "emp_permenent_address", "emp_present_address",
    #                                                                             "emp_status", "emp_hired_date", "emp_active_date", "emp_relegion",
    #                                                                             "emp_blood_group", "emp_nationality_id", "emp_marital_status", "emp_father_name",
    #                                                                             "emp_mother_name", "emp_posting_location", "is_active", "epm_ot_applicable",
    #                                                                             "emp_company_id", "emp_branch_id", "emp_dept_id", "emp_desgntn_id", "emp_ctgry_id"]]
    #             fields_to_include = valid_fields
            
    #         # Create a response object
    #         response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    #         response['Content-Disposition'] = 'attachment; filename="employee_report.xlsx"'

    #         # Create an Excel workbook and add a worksheet
    #         workbook = xlsxwriter.Workbook(response)
    #         worksheet = workbook.add_worksheet()

    #         # Add column headers
    #         headers = ["Employee Code", "First Name", "Last Name", "Gender", "DOB", "Email", "Mobile Number",
    #                 "Mobile Number2", "Country", "State", "City", "Permanent Address", "Present Address", "Status",
    #                 "Hired Date", "Active Date", "Religion", "Blood Group", "Nationality", "Marital Status",
    #                 "Father Name", "Mother Name", "Posting Location", "Active", "OT Applicable", "Company", "Branch",
    #                 "Department", "Designation", "Category"]

    #         # Write column headers for the included fields
    #         for col, field_name in enumerate(fields_to_include):
    #             header = headers[col] if col < len(headers) else field_name  # Use field name if header not available
    #             worksheet.write(0, col, header)
            
    #         worksheet.set_column('A:AD', 15)
            
            # Write data rows
            # date_format = workbook.add_format({'num_format': 'dd-mm-yyyy'})
            # for row, employee in enumerate(employees, start=1):
            #     for col, field_name in enumerate(fields_to_include):
            #         field_value = getattr(employee, field_name)
            #         if field_name == "emp_country_id":
            #             field_value = field_value.country_name if field_value else ""  # Assuming country_name is the attribute
            #         elif field_name == "emp_state_id":
            #             field_value = field_value.state_name if field_value else ""
            #         elif field_name == "emp_nationality_id":
            #             field_value = field_value.N_name if field_value else ""  # Assuming nationality_name is the attribute
            #         elif field_name == "emp_company_id":
            #             field_value = field_value.cmpny_name  if field_value else ""
            #         elif field_name == "emp_branch_id":
            #             field_value = field_value.branch_name if field_value else ""
            #         elif field_name == "emp_dept_id":
            #             field_value = field_value.dept_name if field_value else ""
            #         elif field_name == "emp_desgntn_id": 
            #             field_value = field_value.job_title if field_value else ""
            #         elif field_name == "emp_ctgry_id": 
            #             field_value = field_value.catogary_title if field_value else ""
                    
            #         elif isinstance(field_value, datetime.date):
            #             worksheet.write(row, col, field_value.strftime('%d-%m-%Y'), date_format)
            #         else:
            #             worksheet.write(row, col, field_value)

    #         # Close the workbook
    #         workbook.close()
    #         return response    
    
class CustomFieldViewset(viewsets.ModelViewSet):
    queryset = Emp_CustomField.objects.all()
    serializer_class = CustomFieldSerializer
    permission_classes = [IsAuthenticated]

    def handle_exception(self, exc):
        if isinstance(exc, ValidationError):
            error_messages = [str(error) for error in exc]
            error_message = ', '.join(error_messages)
            return Response({'error': error_message}, status=status.HTTP_400_BAD_REQUEST)
        
        return super().handle_exception(exc)

    


class EmpFam_CustomFieldViewset(viewsets.ModelViewSet):
    queryset = EmpFamily_CustomField.objects.all()
    serializer_class = EmpFam_CustomFieldSerializer
    permission_classes = [IsAuthenticated]

    def handle_exception(self, exc):
        if isinstance(exc, ValidationError):
            error_messages = [str(error) for error in exc]
            error_message = ', '.join(error_messages)
            return Response({'error': error_message}, status=status.HTTP_400_BAD_REQUEST)
        
        return super().handle_exception(exc)


class EmpJobHistory_UdfViewset(viewsets.ModelViewSet):
    queryset = EmpJobHistory_CustomField.objects.all()
    serializer_class = EmpJobHistory_Udf_Serializer
    permission_classes = [IsAuthenticated]

    def handle_exception(self, exc):
        if isinstance(exc, ValidationError):
            error_messages = [str(error) for error in exc]
            error_message = ', '.join(error_messages)
            return Response({'error': error_message}, status=status.HTTP_400_BAD_REQUEST)
        
        return super().handle_exception(exc)


class EmpQf_UdfViewset(viewsets.ModelViewSet):
    queryset = EmpQualification_CustomField.objects.all()
    serializer_class = Emp_qf_udf_Serializer
    permission_classes = [IsAuthenticated]

    def handle_exception(self, exc):
        if isinstance(exc, ValidationError):
            error_messages = [str(error) for error in exc]
            error_message = ', '.join(error_messages)
            return Response({'error': error_message}, status=status.HTTP_400_BAD_REQUEST)
        
        return super().handle_exception(exc)



class EmpDoc_UdfViewset(viewsets.ModelViewSet):
    queryset = EmpDocuments_CustomField.objects.all()
    serializer_class = EmpDocuments_Udf_Serializer
    permission_classes = [IsAuthenticated]

    def handle_exception(self, exc):
        if isinstance(exc, ValidationError):
            error_messages = [str(error) for error in exc]
            error_message = ', '.join(error_messages)
            return Response({'error': error_message}, status=status.HTTP_400_BAD_REQUEST)
        
        return super().handle_exception(exc)



#emp bulkupload

class EmpbulkuploadViewSet(viewsets.ModelViewSet):
    queryset = emp_master.objects.all()
    serializer_class = EmpBulkUploadSerializer
    permission_classes = [IsAuthenticated]
    parser_classes = (MultiPartParser, FormParser)


    @action(detail=False, methods=['post'], parser_classes=[MultiPartParser, FormParser])
    def bulk_upload(self, request):
        if request.method == 'POST' and request.FILES.get('file'):
            excel_file = request.FILES['file']
            if excel_file.name.endswith('.xlsx'):
                try:
                    # Open the Excel file
                    workbook = openpyxl.load_workbook(excel_file)

                    # Initialize error lists
                    all_errors_sheet1 = []
                    all_errors_sheet2 = []

                    # Extract data from each sheet if it exists and is not empty
                    sheet1 = workbook.get_sheet_by_name('Sheet1')
                    sheet2 = workbook.get_sheet_by_name('Sheet2')

                    if sheet1 is None or sheet1.max_row == 1:
                        return Response({"error": "Sheet1 is either missing or empty."}, status=400)

                    if sheet2 is None or sheet2.max_row == 1:
                        return Response({"error": "Sheet2 is either missing or empty."}, status=400)

                    # Convert data to datasets
                    dataset_sheet1 = Dataset()
                    dataset_sheet1.headers = [cell.value for cell in sheet1[1]]
                    for row_idx, row in enumerate(sheet1.iter_rows(min_row=2), start=2):
                        dataset_sheet1.append([cell.value for cell in row])

                    dataset_sheet2 = Dataset()
                    dataset_sheet2.headers = [cell.value for cell in sheet2[1]]
                    for row_idx, row in enumerate(sheet2.iter_rows(min_row=2), start=2):
                        dataset_sheet2.append([str(cell.value) for cell in row])

                    # Create resource instances for Employee and CustomField
                    employee_resource = EmployeeResource()
                    custom_field_resource = EmpCustomFieldResource()

                    # Validate and collect errors from Sheet1
                    with transaction.atomic():
                        for row_idx, row in enumerate(dataset_sheet1.dict, start=2):
                            try:
                                employee_resource.before_import_row(row, row_idx=row_idx)
                            except Exception as e:
                                all_errors_sheet1.append({"row": row_idx, "error": str(e)})

                    # Validate and collect errors from Sheet2
                    with transaction.atomic():
                        for row_idx, row in enumerate(dataset_sheet2.dict, start=2):
                            try:
                                custom_field_resource.before_import_row(row, row_idx=row_idx)
                            except Exception as e:
                                all_errors_sheet2.append({"row": row_idx, "error": str(e)})

                    # Check if there are any errors and return them
                    if all_errors_sheet1 or all_errors_sheet2:
                        return Response({"errors_sheet1": all_errors_sheet1, "errors_sheet2": all_errors_sheet2}, status=400)

                    # If no errors, proceed with importing data

                    # Import data from Sheet1 into emp_master table
                    with transaction.atomic():
                        employee_result = employee_resource.import_data(dataset_sheet1, dry_run=False, raise_errors=True)

                    # Import data from Sheet2 into Emp_CustomField table
                    with transaction.atomic():
                        custom_field_result = custom_field_resource.import_data(dataset_sheet2, dry_run=False, raise_errors=True)

                    return Response({"message": f"{employee_result.total_rows} records created for Sheet1, {custom_field_result.total_rows} records created for Sheet2 successfully"})
                except Exception as e:
                    return Response({"error": str(e)}, status=400)
            else:
                return Response({"error": "Invalid file format. Only Excel files (.xlsx) are supported."}, status=400)
        else:
            return Response({"error": "Please provide an Excel file."}, status=400)
            

    @action(detail=False, methods=['get'])  # New endpoint for downloading default file
    def download_default_excel_file(self, request):
        # demo_file_path = os.path.join(settings.BASE_DIR,'defaults', 'emp mstr.xlsx')
        demo_file_path = os.path.join(os.path.dirname(__file__), 'defaults', 'NewEmpMstr.xlsx')
        try:
            with open(demo_file_path, 'rb') as f:
                response = HttpResponse(f.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = 'attachment; filename="demo.xlsx"'
                return response
        except FileNotFoundError:
            return Response({"error": "Default demo file not found."}, status=400)

#EMP_FAMILY
class EmpFamViewSet(viewsets.ModelViewSet):
    queryset = emp_family.objects.all()  # Retrieve all instances of emp_family model
    serializer_class = EmpFamSerializer  # Use EmpFamSerializer for serialization
    permission_classes = [IsAuthenticated]  # Require authentication for access

    @action(detail=True, methods=['get'])
    def empfamily_udf(self, request, pk=None):
        fam_udf = self.get_object()
        userdefinedfields = fam_udf.custom_fields.all()  
        serializer = EmpFam_CustomFieldSerializer(userdefinedfields, many=True)
        return Response(serializer.data)

    def get_queryset(self):
        user = self.request.user
        if user.is_authenticated:  # Check if user is authenticated
            if user.is_superuser or user.is_staff:  # Check if user is a superuser or staff
                return emp_family.objects.all()  # Return all instances of emp_family
            elif hasattr(user, 'emp_id'):  # If not a superuser or staff, filter based on emp_id
                return emp_family.objects.filter(emp_id=user.emp_id)
            elif user.is_ess:  # If user is an ESS, filter based on created_by
                return emp_family.objects.filter(created_by=user)
        return emp_family.objects.none()  # Return an empty queryset if user is not authenticated or does not meet any condition

#EMP_JOB HISTORY
#EMP_JOB HISTORY
class EmpJobHistoryvSet(viewsets.ModelViewSet):
    queryset = EmpJobHistory.objects.all()
    serializer_class = EmpJobHistorySerializer
    permission_classes = [IsAuthenticated]

    @action(detail=True, methods=['get'])
    def empjobhistory_udf(self, request, pk=None):
        jobhistory_udf = self.get_object()
        userdefinedfields = jobhistory_udf.custom_fields.all()  
        serializer = EmpJobHistory_Udf_Serializer(userdefinedfields, many=True)
        return Response(serializer.data)

    def get_queryset(self):
        user = self.request.user
        if user.is_authenticated:
            if user.is_superuser or user.is_staff:  # Check if user is a superuser or staff
                return EmpJobHistory.objects.all()  # Return all instances of emp_family
            elif hasattr(user, 'emp_id'):  # Assuming 'emp_id' is the attribute that stores employee ID
                return EmpJobHistory.objects.filter(emp_id=user.emp_id)
            elif user.is_ess:  # If user is an ESS, filter based on created_by
                return EmpJobHistory.objects.filter(created_by=user)
        return EmpJobHistory.objects.none()
    def get_serializer_context(self):
        return {'request': self.request}
    
#EMP_QUALIFICATION HISTORY


class Emp_QualificationViewSet(viewsets.ModelViewSet):
    queryset = EmpQualification.objects.all()
    serializer_class = Emp_qf_Serializer
    permission_classes = [IsAuthenticated]

    @action(detail=True, methods=['get'])
    def empqualification_udf(self, request, pk=None):
        emp_qf_udf = self.get_object()
        userdefinedfields = emp_qf_udf.custom_fields.all()  
        serializer = Emp_qf_Serializer(userdefinedfields, many=True)
        return Response(serializer.data)

    def get_queryset(self):
        user = self.request.user
        if user.is_authenticated:
            if user.is_superuser or user.is_staff:  # Check if user is a superuser or staff
                return EmpQualification.objects.all()  # Return all instances of emp_family
            elif hasattr(user, 'emp_id'):  # Assuming 'emp_id' is the attribute that stores employee ID
                return EmpQualification.objects.filter(emp_id=user.emp_id)
            elif user.is_ess:  # If user is an ESS, filter based on created_by
                return EmpQualification.objects.filter(created_by=user)
        return EmpQualification.objects.none()
    def get_serializer_context(self):
        return {'request': self.request}
    

#EMP_DOCUMENT 



class Emp_DocumentViewSet(viewsets.ModelViewSet):
    queryset = Emp_Documents.objects.all()
    serializer_class = DocumentSerializer
    permission_classes = [IsAuthenticated]
    
    
    @action(detail=True, methods=['get'])
    def empfamily_udf(self, request, pk=None):
        empDoc_udf = self.get_object()
        userdefinedfields = empDoc_udf.custom_fields.all()  
        serializer = EmpDocuments_Udf_Serializer(userdefinedfields, many=True)
        return Response(serializer.data)
    
    
    def get_serializer_context(self):
        return {'request': self.request}
    def get_queryset(self):
        user = self.request.user
        if user.is_authenticated:
            if user.is_superuser or user.is_staff:  # Check if user is a superuser or staff
                return Emp_Documents.objects.all()  # Return all instances of emp_family
            elif hasattr(user, 'emp_id'):  # Assuming 'emp_id' is the attribute that stores employee ID
                return Emp_Documents.objects.filter(emp_id=user.emp_id)
            elif user.is_ess:  # If user is an ESS, filter based on created_by
                return Emp_Documents.objects.filter(created_by=user)
        return Emp_Documents.objects.none()
        # if user.emp_doc_expiry_date - date.today().days <= 7

    
class Bulkupload_DocumentViewSet(viewsets.ModelViewSet):
    queryset = Emp_Documents.objects.all()
    serializer_class = DocBulkuploadSerializer
    permission_classes = [IsAuthenticated]
    parser_classes = (MultiPartParser, FormParser)
    
    @action(detail=False, methods=['post'], parser_classes=[MultiPartParser, FormParser])
    def bulk_upload(self, request):
        if request.method == 'POST' and request.FILES.get('file'):
            excel_file = request.FILES['file']
            if excel_file.name.endswith('.xlsx'):
                try:
                    # Load data from the Excel file into a Dataset
                    dataset = Dataset()
                    dataset.load(excel_file.read(), format='xlsx')

                    # Create a resource instance
                    resource = DocumentResource()

                    # Analyze the entire dataset to collect errors
                    all_errors = []
                    valid_rows = []
                    with transaction.atomic():
                        for row_idx, row in enumerate(dataset.dict, start=2):  # Start from row 2 (1-based index)
                            row_errors = []
                            try:
                                resource.before_import_row(row, row_idx=row_idx)
                            except ValidationError as e:
                                row_errors.extend([f"Row {row_idx}: {error}" for error in e.messages])
                            if row_errors:
                                all_errors.extend(row_errors)
                            else:
                                valid_rows.append(row)

                        # After analyzing all rows, check for duplicate values
                        # duplicate_errors = resource.check_duplicate_values(dataset)

                        # # If there are any duplicate errors, add them to the list of all errors
                        # if duplicate_errors:
                        #     all_errors.extend(duplicate_errors)

                    # If there are any errors, return them
                    if all_errors:
                        return Response({"errors": all_errors}, status=400)

                    # If no errors, import valid data into the model
                    with transaction.atomic():
                        result = resource.import_data(dataset, dry_run=False, raise_errors=True)

                    return Response({"message": f"{result.total_rows} records created successfully"})
                except Exception as e:
                    return Response({"error": str(e)}, status=400)
            else:
                return Response({"error": "Invalid file format. Only Excel files (.xlsx) are supported."}, status=400)
        else:
            return Response({"error": "Please provide an Excel file."}, status=400)




# EmpLeaveRequest
class EmpLeaveRequestViewSet(viewsets.ModelViewSet):
    queryset = EmpLeaveRequest.objects.all()
    serializer_class = EmpLeaveRequestSerializer
    permission_classes = [IsAuthenticated]
    def get_serializer_context(self):
        return {'request': self.request}
    


class LanguageSkillViewSet(viewsets.ModelViewSet):
    queryset = LanguageSkill.objects.all()
    serializer_class = LanguageSkillSerializer

class MarketingSkillViewSet(viewsets.ModelViewSet):
    queryset = MarketingSkill.objects.all()
    serializer_class = MarketingSkillSerializer

class ProgrammingLanguageSkillViewSet(viewsets.ModelViewSet):
    queryset = ProgrammingLanguageSkill.objects.all()
    serializer_class = ProgrammingLanguageSkillSerializer

class EmployeeSkillViewSet(viewsets.ModelViewSet):
    queryset = EmployeeSkill.objects.all()
    serializer_class = EmployeeSkillSerializer

class LanguageBlkupldViewSet(viewsets.ModelViewSet):
    queryset = LanguageSkill.objects.all()
    serializer_class = LanguageBlkupldSerializer
    permission_classes = [IsAuthenticated]
    parser_classes = (MultiPartParser, FormParser)

    @action(detail=False, methods=['post'], parser_classes=[MultiPartParser, FormParser])
    def bulk_upload(self, request):
        if request.method == 'POST' and request.FILES.get('file'):
            excel_file = request.FILES['file']
            
            # Check if the uploaded file is an Excel file
            if not excel_file.name.endswith('.xlsx'):
                return Response({'error': 'Only Excel files (.xlsx) are supported'}, status=status.HTTP_400_BAD_REQUEST)
            
            try:
                # Read the Excel file using pandas
                df = pd.read_excel(excel_file)
                
                # Iterate through each row in the DataFrame
                for index, row in df.iterrows():
                    # Get the emp_master instance corresponding to the emp_id
                    
                    
                    # Create a Skills_Master object with the emp_instance
                    LanguageSkill.objects.create(
                        
                        language=row['language'],
                        
                    )
            except Exception as e:
                return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

            return Response({'message': 'Bulk upload successful'}, status=status.HTTP_201_CREATED)
        else:
            return Response({'error': 'No file found'}, status=status.HTTP_400_BAD_REQUEST)

class MarketingBlkupldViewSet(viewsets.ModelViewSet):
    queryset = MarketingSkill.objects.all()
    serializer_class = MarketingBlkupldSerializer
    permission_classes = [IsAuthenticated]
    parser_classes = (MultiPartParser, FormParser)

    @action(detail=False, methods=['post'], parser_classes=[MultiPartParser, FormParser])
    def bulk_upload(self, request):
        if request.method == 'POST' and request.FILES.get('file'):
            excel_file = request.FILES['file']
            
            # Check if the uploaded file is an Excel file
            if not excel_file.name.endswith('.xlsx'):
                return Response({'error': 'Only Excel files (.xlsx) are supported'}, status=status.HTTP_400_BAD_REQUEST)
            
            try:
                # Read the Excel file using pandas
                df = pd.read_excel(excel_file)
                
                # Iterate through each row in the DataFrame
                for index, row in df.iterrows():
                    # Get the emp_master instance corresponding to the emp_id
                    
                    
                    # Create a Skills_Master object with the emp_instance
                    MarketingSkill.objects.create(
                        
                        marketing=row['marketing'],
                        
                    )
            except Exception as e:
                return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

            return Response({'message': 'Bulk upload successful'}, status=status.HTTP_201_CREATED)
        else:
            return Response({'error': 'No file found'}, status=status.HTTP_400_BAD_REQUEST)


class ProLangBlkupldViewSet(viewsets.ModelViewSet):
    queryset = ProgrammingLanguageSkill.objects.all()
    serializer_class = ProLangBlkupldSerializer
    permission_classes = [IsAuthenticated]
    parser_classes = (MultiPartParser, FormParser)

    @action(detail=False, methods=['post'], parser_classes=[MultiPartParser, FormParser])
    def bulk_upload(self, request):
        if request.method == 'POST' and request.FILES.get('file'):
            excel_file = request.FILES['file']
            
            # Check if the uploaded file is an Excel file
            if not excel_file.name.endswith('.xlsx'):
                return Response({'error': 'Only Excel files (.xlsx) are supported'}, status=status.HTTP_400_BAD_REQUEST)
            
            try:
                # Read the Excel file using pandas
                df = pd.read_excel(excel_file)
                
                # Iterate through each row in the DataFrame
                for index, row in df.iterrows():
                    # Get the emp_master instance corresponding to the emp_id
                    
                    
                    # Create a Skills_Master object with the emp_instance
                    ProgrammingLanguageSkill.objects.create(
                        
                        programming_language=row['programming_language'],
                        
                    )
            except Exception as e:
                return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

            return Response({'message': 'Bulk upload successful'}, status=status.HTTP_201_CREATED)
        else:
            return Response({'error': 'No file found'}, status=status.HTTP_400_BAD_REQUEST)




# def create_employee_skill(request):
#     if request.method == 'POST':
#         form = EmployeeSkillForm(request.POST)
#         if form.is_valid():
#             form.save()
#             # Redirect or do something else on successful form submission
#     else:
#         form = EmployeeSkillForm()
#     return render(request, 'your_template.html', {'form': form})

# class SkillMasterViewSet(viewsets.ModelViewSet):
#     queryset = Skills_Master.objects.all()
#     serializer_class = SkillMasterSerializer

# class SkillsBlkupldViewSet(viewsets.ModelViewSet):
#     queryset = Skills_Master.objects.all()
#     serializer_class = SkillsBlkupldSerializer
#     permission_classes = [IsAuthenticated]
#     parser_classes = (MultiPartParser, FormParser)

#     @action(detail=False, methods=['post'], parser_classes=[MultiPartParser, FormParser])
#     def bulk_upload(self, request):
#         if request.method == 'POST' and request.FILES.get('file'):
#             excel_file = request.FILES['file']
            
#             # Check if the uploaded file is an Excel file
#             if not excel_file.name.endswith('.xlsx'):
#                 return Response({'error': 'Only Excel files (.xlsx) are supported'}, status=status.HTTP_400_BAD_REQUEST)
            
#             try:
#                 # Read the Excel file using pandas
#                 df = pd.read_excel(excel_file)
                
#                 # Iterate through each row in the DataFrame
#                 for index, row in df.iterrows():
#                     # Get the emp_master instance corresponding to the emp_id
#                     emp_instance = emp_master.objects.get(id=row['emp_id'])
                    
#                     # Create a Skills_Master object with the emp_instance
#                     Skills_Master.objects.create(
#                         emp_id=emp_instance,
#                         language=row['language'],
#                         marketing=row['marketing'],
#                         programming_language=row['programming_language']
#                     )
#             except Exception as e:
#                 return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

#             return Response({'message': 'Bulk upload successful'}, status=status.HTTP_201_CREATED)
#         else:
#             return Response({'error': 'No file found'}, status=status.HTTP_400_BAD_REQUEST)